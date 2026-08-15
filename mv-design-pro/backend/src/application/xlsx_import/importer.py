"""
Import sieci SN z arkusza XLSX — ODCZYT I WALIDACJA (bez zapisu do bazy).

Format arkusza (uzgodniony z operatorami sieci):
- Arkusz "Szyny"    (wymagany):  id, nazwa, napięcie_kV
- Arkusz "Linie"    (wymagany):  id, szyna_pocz, szyna_kon, typ, długość_km,
                                 R_ohm_km, X_ohm_km [opcjonalnie: B_uS_km, typ_katalogowy]
- Arkusz "Trafo"    (opcjonalny): id, szyna_HV, szyna_LV, Sn_MVA, uk_pct [opc.: Pk_kW, grupa]
- Arkusz "Źródła"   (opcjonalny): id, szyna, typ, Sk_MVA, RX_ratio
- Arkusz "Odbiory"  (opcjonalny): id, szyna, P_MW, Q_Mvar

ZASADY (naprawa karty XLSX-IMPORT, 2026-08-07):
- ZERO FIZYKI. Ten moduł NIE liczy zadnej wielkosci elektrycznej. Dane zrodla
  (Sk'', R/X) trafiaja do modelu jako DANE WEJSCIOWE w kanonicznym ksztalcie
  (`model`/`sk3_mva`/`rx_ratio` — jak w kreatorze sieci); impedancje zastepcza
  liczy warstwa solverowa. Stan PRZED: importer sam liczyl Z = Un²/Sk'' i R/X,
  po czym wpisywal wynik w DYNAMICZNY atrybut `node.source_impedance`, ktorego
  NIKT w systemie nie czytal (fizyka w warstwie aplikacji + martwa dana).
- ZERO ZGADYWANIA. Nazwa handlowa z arkusza (kolumna `typ`) NIE jest dopasowywana
  do katalogu po podobienstwie — wiazanie katalogowe powstaje wylacznie z jawnej
  kolumny `typ_katalogowy` (identyfikator typu). Brak wiazania nie jest bledem:
  element trafia do BRAMKI KATALOGOWEJ (jak przy imporcie archiwum ZIP), a
  projektant domapowuje typ w katalogu.
- ZERO WARTOSCI FIKCYJNYCH. Stan PRZED wpisywal `rated_current_a=1.0` z komentarzem
  „placeholder"; teraz obciazalnosc pochodzi WYLACZNIE z typu katalogowego, a jej
  brak zostaje brakiem (0.0 = wielkosc nieznana, kryterium niesprawdzalne).
- Bledy sa strukturalne (arkusz/wiersz/kolumna/komunikat) — front pokazuje je
  przy wierszu, bez parsowania arkusza w przegladarce.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from typing import Any

from network_model.catalog.governance import wymaga_referencji_katalogowej
from network_model.catalog.repository import CatalogRepository, get_default_mv_catalog
from network_model.core.branch import BranchType

# Nazwy arkuszy i kolumn — jedyne zrodlo prawdy formatu (uzywane tez przez API/dokumentacje).
ARKUSZ_SZYNY = "Szyny"
ARKUSZ_LINIE = "Linie"
ARKUSZ_TRAFO = "Trafo"
ARKUSZ_ZRODLA = "Źródła"
ARKUSZ_ODBIORY = "Odbiory"

ARKUSZE_WYMAGANE: tuple[str, ...] = (ARKUSZ_SZYNY, ARKUSZ_LINIE)
ARKUSZE_OPCJONALNE: tuple[str, ...] = (ARKUSZ_TRAFO, ARKUSZ_ZRODLA, ARKUSZ_ODBIORY)


class XlsxValidationError(Exception):
    """Blad walidacji importu XLSX."""

    def __init__(self, errors: list[str]):
        self.errors = errors
        super().__init__(f"Błędy walidacji importu XLSX: {len(errors)} błędów")


@dataclass(frozen=True)
class BladArkusza:
    """Pojedyncze zastrzezenie do zawartosci arkusza.

    `wiersz` = None oznacza zastrzezenie do calego arkusza (np. brak kolumny),
    `kolumna` = None oznacza zastrzezenie do calego wiersza (np. brak szyny docelowej).
    """

    arkusz: str
    komunikat: str
    wiersz: int | None = None
    kolumna: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "arkusz": self.arkusz,
            "wiersz": self.wiersz,
            "kolumna": self.kolumna,
            "komunikat": self.komunikat,
        }

    def jako_tekst(self) -> str:
        czesci = [f"Arkusz '{self.arkusz}'"]
        if self.wiersz is not None:
            czesci.append(f"wiersz {self.wiersz}")
        if self.kolumna is not None:
            czesci.append(f"kolumna '{self.kolumna}'")
        return f"{', '.join(czesci)}: {self.komunikat}"


@dataclass(frozen=True)
class SiecZArkusza:
    """Zawartosc arkusza przelozona na kanoniczne rekordy modelu (bez zapisu).

    Rekordy maja ksztalt kanoniczny warstwy trwalosci (`network_nodes`,
    `network_branches`, `network_sources`, `network_loads`), zeby zapis byl
    czystym przepisaniem, bez drugiego tlumaczenia danych.
    """

    wezly: list[dict[str, Any]] = field(default_factory=list)
    galezie: list[dict[str, Any]] = field(default_factory=list)
    zrodla: list[dict[str, Any]] = field(default_factory=list)
    odbiory: list[dict[str, Any]] = field(default_factory=list)


@dataclass
class XlsxImportResult:
    """Wynik odczytu arkusza (bez zapisu do bazy)."""

    success: bool
    siec: SiecZArkusza | None = None
    bus_count: int = 0
    branch_count: int = 0
    source_count: int = 0
    load_count: int = 0
    trafo_count: int = 0
    warnings: list[str] = field(default_factory=list)
    bledy: list[BladArkusza] = field(default_factory=list)
    elementy_bez_katalogu: list[str] = field(default_factory=list)
    mapowanie_katalogowe_wymagane: bool = False

    @property
    def errors(self) -> list[str]:
        """Bledy jako plaskie komunikaty (zgodnosc z dotychczasowym kontraktem API)."""
        return [blad.jako_tekst() for blad in self.bledy]

    def to_dict(self) -> dict[str, Any]:
        return {
            "success": self.success,
            "bus_count": self.bus_count,
            "branch_count": self.branch_count,
            "source_count": self.source_count,
            "load_count": self.load_count,
            "trafo_count": self.trafo_count,
            "warnings": self.warnings,
            "errors": self.errors,
            "bledy": [blad.to_dict() for blad in self.bledy],
            "elementy_bez_katalogu": self.elementy_bez_katalogu,
            "mapowanie_katalogowe_wymagane": self.mapowanie_katalogowe_wymagane,
        }


class XlsxNetworkImporter:
    """Odczyt sieci SN z arkusza Excel.

    Uzycie:
        importer = XlsxNetworkImporter()
        wynik = importer.import_from_bytes(xlsx_bytes)
        if wynik.success:
            siec = wynik.siec  # rekordy do zapisu przez XlsxImportService
    """

    REQUIRED_SHEETS = set(ARKUSZE_WYMAGANE)
    OPTIONAL_SHEETS = set(ARKUSZE_OPCJONALNE)

    BUS_COLUMNS: dict[str, type] = {"id": str, "nazwa": str, "napięcie_kV": float}
    LINE_COLUMNS: dict[str, type] = {
        "id": str,
        "szyna_pocz": str,
        "szyna_kon": str,
        "typ": str,
        "długość_km": float,
        "R_ohm_km": float,
        "X_ohm_km": float,
    }
    LINE_OPTIONAL_COLUMNS: dict[str, type] = {"B_uS_km": float, "typ_katalogowy": str}
    TRAFO_COLUMNS: dict[str, type] = {
        "id": str,
        "szyna_HV": str,
        "szyna_LV": str,
        "Sn_MVA": float,
        "uk_pct": float,
    }
    TRAFO_OPTIONAL_COLUMNS: dict[str, type] = {"Pk_kW": float, "grupa": str}
    SOURCE_COLUMNS: dict[str, type] = {
        "id": str,
        "szyna": str,
        "typ": str,
        "Sk_MVA": float,
        "RX_ratio": float,
    }
    LOAD_COLUMNS: dict[str, type] = {
        "id": str,
        "szyna": str,
        "P_MW": float,
        "Q_Mvar": float,
    }

    def __init__(self, catalog: CatalogRepository | None = None) -> None:
        self._catalog = catalog

    def _katalog(self) -> CatalogRepository:
        if self._catalog is None:
            self._catalog = get_default_mv_catalog()
        return self._catalog

    # ------------------------------------------------------------------
    # Wejscie glowne
    # ------------------------------------------------------------------

    def import_from_bytes(self, data: bytes) -> XlsxImportResult:
        """Odczytaj siec z bajtow pliku XLSX."""
        import openpyxl  # zaleznosc glowna (pyproject) — brak = blad srodowiska, nie danych

        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        except Exception as e:
            return XlsxImportResult(
                success=False,
                bledy=[
                    BladArkusza(
                        arkusz="—",
                        komunikat=(
                            "Nie można otworzyć pliku jako arkusza XLSX "
                            f"(plik uszkodzony lub w innym formacie): {e}"
                        ),
                    )
                ],
            )

        try:
            bledy: list[BladArkusza] = []
            ostrzezenia: list[str] = []

            nazwy_arkuszy = set(wb.sheetnames)
            for wymagany in ARKUSZE_WYMAGANE:
                if wymagany not in nazwy_arkuszy:
                    bledy.append(
                        BladArkusza(
                            arkusz=wymagany,
                            komunikat="Brak wymaganego arkusza w pliku",
                        )
                    )
            if bledy:
                return XlsxImportResult(success=False, bledy=bledy)

            # RAPORT ZBIORCZY: parsowanie NIE przerywa pracy przy pierwszym potknieciu.
            # Wiersz z bledna komorka jest pomijany, ale pozostale wiersze i pozostale
            # arkusze sa czytane dalej, a walidacje krzyzowe biegna na tym, co sie
            # odczytalo. Inaczej projektant poprawia arkusz w ping-pongu: jeden blad na
            # wyslanie. Przerwanie zostaje TYLKO dla braku arkusza/kolumny — bez kolumny
            # nie ma czego walidowac.
            struktura_ok = True
            szyny, ok = self._parse_sheet(
                wb[ARKUSZ_SZYNY], self.BUS_COLUMNS, {}, ARKUSZ_SZYNY, bledy
            )
            struktura_ok &= ok
            linie, ok = self._parse_sheet(
                wb[ARKUSZ_LINIE],
                self.LINE_COLUMNS,
                self.LINE_OPTIONAL_COLUMNS,
                ARKUSZ_LINIE,
                bledy,
            )
            struktura_ok &= ok
            trafo: list[dict[str, Any]] = []
            if ARKUSZ_TRAFO in nazwy_arkuszy:
                trafo, ok = self._parse_sheet(
                    wb[ARKUSZ_TRAFO],
                    self.TRAFO_COLUMNS,
                    self.TRAFO_OPTIONAL_COLUMNS,
                    ARKUSZ_TRAFO,
                    bledy,
                )
                struktura_ok &= ok
            zrodla: list[dict[str, Any]] = []
            if ARKUSZ_ZRODLA in nazwy_arkuszy:
                zrodla, ok = self._parse_sheet(
                    wb[ARKUSZ_ZRODLA], self.SOURCE_COLUMNS, {}, ARKUSZ_ZRODLA, bledy
                )
                struktura_ok &= ok
            odbiory: list[dict[str, Any]] = []
            if ARKUSZ_ODBIORY in nazwy_arkuszy:
                odbiory, ok = self._parse_sheet(
                    wb[ARKUSZ_ODBIORY], self.LOAD_COLUMNS, {}, ARKUSZ_ODBIORY, bledy
                )
                struktura_ok &= ok
        finally:
            wb.close()

        if not struktura_ok:
            return XlsxImportResult(success=False, bledy=bledy)

        bledy.extend(self._waliduj_powiazania(szyny, linie, trafo, zrodla, odbiory))
        bledy.extend(self._waliduj_wartosci(szyny, linie, trafo, zrodla))
        bledy.extend(self._waliduj_typy_katalogowe(linie))

        if bledy:
            return XlsxImportResult(success=False, bledy=bledy)

        siec, elementy_bez_katalogu = self._zbuduj_rekordy(
            szyny, linie, trafo, zrodla, odbiory, ostrzezenia
        )

        return XlsxImportResult(
            success=True,
            siec=siec,
            bus_count=len(szyny),
            branch_count=len(linie) + len(trafo),
            source_count=len(zrodla),
            load_count=len(odbiory),
            trafo_count=len(trafo),
            warnings=ostrzezenia,
            elementy_bez_katalogu=elementy_bez_katalogu,
            mapowanie_katalogowe_wymagane=bool(elementy_bez_katalogu),
        )

    # ------------------------------------------------------------------
    # Odczyt arkusza
    # ------------------------------------------------------------------

    def _parse_sheet(
        self,
        sheet: Any,
        kolumny: dict[str, type],
        kolumny_opcjonalne: dict[str, type],
        nazwa_arkusza: str,
        bledy: list[BladArkusza],
    ) -> tuple[list[dict[str, Any]], bool]:
        """Odczytaj arkusz do listy rekordow.

        Returns:
            (rekordy poprawnych wierszy, czy STRUKTURA arkusza jest czytelna).
            Struktura nieczytelna (pusty arkusz, brak kolumny) => dalsze walidacje
            krzyzowe nie maja sensu i wolajacy przerywa.
        """
        wiersze = list(sheet.iter_rows(min_row=1, values_only=True))
        if not wiersze:
            bledy.append(BladArkusza(arkusz=nazwa_arkusza, komunikat="Arkusz jest pusty"))
            return [], False

        naglowki = [str(h).strip() if h is not None else "" for h in wiersze[0]]

        brakujace = [nazwa for nazwa in kolumny if nazwa not in naglowki]
        for nazwa in brakujace:
            bledy.append(
                BladArkusza(
                    arkusz=nazwa_arkusza,
                    kolumna=nazwa,
                    komunikat="Brak wymaganej kolumny w nagłówku arkusza",
                )
            )
        if brakujace:
            return [], False

        indeksy = {nazwa: naglowki.index(nazwa) for nazwa in kolumny}
        indeksy_opcjonalne = {
            nazwa: naglowki.index(nazwa) for nazwa in kolumny_opcjonalne if nazwa in naglowki
        }

        wiersze_danych = wiersze[1:]
        if not any(self._wiersz_niepusty(w) for w in wiersze_danych):
            bledy.append(
                BladArkusza(
                    arkusz=nazwa_arkusza,
                    komunikat="Arkusz nie zawiera żadnego wiersza danych (sam nagłówek)",
                )
            )
            return [], False

        rekordy: list[dict[str, Any]] = []
        for numer_wiersza, wiersz in enumerate(wiersze_danych, start=2):
            if not self._wiersz_niepusty(wiersz):
                continue  # pusty wiersz separujacy — pomijany, nie jest bledem danych

            rekord: dict[str, Any] = {}
            wiersz_poprawny = True

            for nazwa, typ_kolumny in kolumny.items():
                wartosc = self._komorka(wiersz, indeksy[nazwa])
                if wartosc is None:
                    bledy.append(
                        BladArkusza(
                            arkusz=nazwa_arkusza,
                            wiersz=numer_wiersza,
                            kolumna=nazwa,
                            komunikat="Pusta wartość w wymaganej kolumnie",
                        )
                    )
                    wiersz_poprawny = False
                    continue
                przekonwertowana = self._konwertuj(wartosc, typ_kolumny)
                if przekonwertowana is None:
                    bledy.append(
                        BladArkusza(
                            arkusz=nazwa_arkusza,
                            wiersz=numer_wiersza,
                            kolumna=nazwa,
                            komunikat=(
                                f"Nieprawidłowa wartość '{wartosc}' — "
                                f"oczekiwano: {self._nazwa_typu(typ_kolumny)}"
                            ),
                        )
                    )
                    wiersz_poprawny = False
                    continue
                rekord[nazwa] = przekonwertowana

            for nazwa, typ_kolumny in kolumny_opcjonalne.items():
                if nazwa not in indeksy_opcjonalne:
                    continue
                wartosc = self._komorka(wiersz, indeksy_opcjonalne[nazwa])
                if wartosc is None:
                    continue
                przekonwertowana = self._konwertuj(wartosc, typ_kolumny)
                if przekonwertowana is None:
                    bledy.append(
                        BladArkusza(
                            arkusz=nazwa_arkusza,
                            wiersz=numer_wiersza,
                            kolumna=nazwa,
                            komunikat=(
                                f"Nieprawidłowa wartość '{wartosc}' — "
                                f"oczekiwano: {self._nazwa_typu(typ_kolumny)}"
                            ),
                        )
                    )
                    wiersz_poprawny = False
                    continue
                rekord[nazwa] = przekonwertowana

            rekord["_wiersz"] = numer_wiersza
            if wiersz_poprawny:
                rekordy.append(rekord)

        return rekordy, True

    @staticmethod
    def _wiersz_niepusty(wiersz: tuple[Any, ...]) -> bool:
        return any(komorka is not None and str(komorka).strip() != "" for komorka in wiersz)

    @staticmethod
    def _komorka(wiersz: tuple[Any, ...], indeks: int) -> Any:
        wartosc = wiersz[indeks] if indeks < len(wiersz) else None
        if wartosc is None or str(wartosc).strip() == "":
            return None
        return wartosc

    @staticmethod
    def _konwertuj(wartosc: Any, typ_kolumny: type) -> Any:
        try:
            if typ_kolumny is str:
                return str(wartosc).strip()
            return typ_kolumny(wartosc)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _nazwa_typu(typ_kolumny: type) -> str:
        return "liczba" if typ_kolumny is float else "tekst"

    # ------------------------------------------------------------------
    # Walidacje
    # ------------------------------------------------------------------

    def _waliduj_powiazania(
        self,
        szyny: list[dict[str, Any]],
        linie: list[dict[str, Any]],
        trafo: list[dict[str, Any]],
        zrodla: list[dict[str, Any]],
        odbiory: list[dict[str, Any]],
    ) -> list[BladArkusza]:
        bledy: list[BladArkusza] = []
        identyfikatory_szyn = {s["id"] for s in szyny}

        odwolania = (
            (ARKUSZ_LINIE, linie, ("szyna_pocz", "szyna_kon")),
            (ARKUSZ_TRAFO, trafo, ("szyna_HV", "szyna_LV")),
            (ARKUSZ_ZRODLA, zrodla, ("szyna",)),
            (ARKUSZ_ODBIORY, odbiory, ("szyna",)),
        )
        for nazwa_arkusza, rekordy, kolumny in odwolania:
            for rekord in rekordy:
                for kolumna in kolumny:
                    if rekord[kolumna] not in identyfikatory_szyn:
                        bledy.append(
                            BladArkusza(
                                arkusz=nazwa_arkusza,
                                wiersz=rekord["_wiersz"],
                                kolumna=kolumna,
                                komunikat=(
                                    f"Szyna '{rekord[kolumna]}' nie występuje "
                                    f"w arkuszu '{ARKUSZ_SZYNY}'"
                                ),
                            )
                        )

        duplikaty = (
            (ARKUSZ_SZYNY, szyny),
            (ARKUSZ_LINIE, linie),
            (ARKUSZ_TRAFO, trafo),
            (ARKUSZ_ZRODLA, zrodla),
            (ARKUSZ_ODBIORY, odbiory),
        )
        for nazwa_arkusza, rekordy in duplikaty:
            widziane: dict[str, int] = {}
            for rekord in rekordy:
                identyfikator = rekord["id"]
                if identyfikator in widziane:
                    bledy.append(
                        BladArkusza(
                            arkusz=nazwa_arkusza,
                            wiersz=rekord["_wiersz"],
                            kolumna="id",
                            komunikat=(
                                f"Powtórzony identyfikator '{identyfikator}' "
                                f"(pierwsze wystąpienie: wiersz {widziane[identyfikator]})"
                            ),
                        )
                    )
                else:
                    widziane[identyfikator] = rekord["_wiersz"]

        return bledy

    def _waliduj_wartosci(
        self,
        szyny: list[dict[str, Any]],
        linie: list[dict[str, Any]],
        trafo: list[dict[str, Any]],
        zrodla: list[dict[str, Any]],
    ) -> list[BladArkusza]:
        bledy: list[BladArkusza] = []

        for szyna in szyny:
            if szyna["napięcie_kV"] <= 0:
                bledy.append(
                    BladArkusza(
                        arkusz=ARKUSZ_SZYNY,
                        wiersz=szyna["_wiersz"],
                        kolumna="napięcie_kV",
                        komunikat="Napięcie znamionowe musi być większe od zera",
                    )
                )

        for linia in linie:
            if linia["długość_km"] <= 0:
                bledy.append(
                    BladArkusza(
                        arkusz=ARKUSZ_LINIE,
                        wiersz=linia["_wiersz"],
                        kolumna="długość_km",
                        komunikat="Długość musi być większa od zera",
                    )
                )
            for kolumna in ("R_ohm_km", "X_ohm_km"):
                if linia[kolumna] < 0:
                    bledy.append(
                        BladArkusza(
                            arkusz=ARKUSZ_LINIE,
                            wiersz=linia["_wiersz"],
                            kolumna=kolumna,
                            komunikat="Wartość jednostkowa nie może być ujemna",
                        )
                    )
            if linia["szyna_pocz"] == linia["szyna_kon"]:
                bledy.append(
                    BladArkusza(
                        arkusz=ARKUSZ_LINIE,
                        wiersz=linia["_wiersz"],
                        kolumna="szyna_kon",
                        komunikat="Początek i koniec odcinka wskazują tę samą szynę",
                    )
                )

        for transformator in trafo:
            if transformator["Sn_MVA"] <= 0:
                bledy.append(
                    BladArkusza(
                        arkusz=ARKUSZ_TRAFO,
                        wiersz=transformator["_wiersz"],
                        kolumna="Sn_MVA",
                        komunikat="Moc znamionowa musi być większa od zera",
                    )
                )
            if transformator["uk_pct"] <= 0:
                bledy.append(
                    BladArkusza(
                        arkusz=ARKUSZ_TRAFO,
                        wiersz=transformator["_wiersz"],
                        kolumna="uk_pct",
                        komunikat="Napięcie zwarcia musi być większe od zera",
                    )
                )

        for zrodlo in zrodla:
            if zrodlo["Sk_MVA"] <= 0:
                bledy.append(
                    BladArkusza(
                        arkusz=ARKUSZ_ZRODLA,
                        wiersz=zrodlo["_wiersz"],
                        kolumna="Sk_MVA",
                        komunikat="Moc zwarciowa musi być większa od zera",
                    )
                )
            if zrodlo["RX_ratio"] < 0:
                bledy.append(
                    BladArkusza(
                        arkusz=ARKUSZ_ZRODLA,
                        wiersz=zrodlo["_wiersz"],
                        kolumna="RX_ratio",
                        komunikat="Stosunek R/X nie może być ujemny",
                    )
                )

        return bledy

    def _waliduj_typy_katalogowe(self, linie: list[dict[str, Any]]) -> list[BladArkusza]:
        """Jawnie podany typ katalogowy MUSI istnieć w katalogu (żadnego cichego pominięcia)."""
        bledy: list[BladArkusza] = []
        katalog = self._katalog()
        for linia in linie:
            typ_katalogowy = linia.get("typ_katalogowy")
            if not typ_katalogowy:
                continue
            if typ_katalogowy in katalog.line_types or typ_katalogowy in katalog.cable_types:
                continue
            bledy.append(
                BladArkusza(
                    arkusz=ARKUSZ_LINIE,
                    wiersz=linia["_wiersz"],
                    kolumna="typ_katalogowy",
                    komunikat=(
                        f"Typ '{typ_katalogowy}' nie występuje w katalogu — "
                        f"popraw identyfikator albo usuń kolumnę i domapuj typ po imporcie"
                    ),
                )
            )
        return bledy

    # ------------------------------------------------------------------
    # Budowa rekordow kanonicznych
    # ------------------------------------------------------------------

    def _zbuduj_rekordy(
        self,
        szyny: list[dict[str, Any]],
        linie: list[dict[str, Any]],
        trafo: list[dict[str, Any]],
        zrodla: list[dict[str, Any]],
        odbiory: list[dict[str, Any]],
        ostrzezenia: list[str],
    ) -> tuple[SiecZArkusza, list[str]]:
        katalog = self._katalog()
        szyny_ze_zrodlem = {z["szyna"] for z in zrodla}

        wezly: list[dict[str, Any]] = []
        slack_przypisany = False
        for szyna in szyny:
            czy_slack = szyna["id"] in szyny_ze_zrodlem and not slack_przypisany
            if czy_slack:
                slack_przypisany = True
                wezly.append(
                    {
                        "ref": szyna["id"],
                        "name": szyna["nazwa"],
                        "node_type": "SLACK",
                        "base_kv": szyna["napięcie_kV"],
                        "attrs": {"voltage_magnitude_pu": 1.0, "voltage_angle_rad": 0.0},
                    }
                )
            else:
                wezly.append(
                    {
                        "ref": szyna["id"],
                        "name": szyna["nazwa"],
                        "node_type": "PQ",
                        "base_kv": szyna["napięcie_kV"],
                        "attrs": {},
                    }
                )

        if zrodla and not slack_przypisany:
            ostrzezenia.append(
                "Arkusz nie pozwolił wskazać szyny bilansującej — sprawdź arkusz 'Źródła'."
            )

        elementy_bez_katalogu: list[str] = []
        galezie: list[dict[str, Any]] = []

        for linia in linie:
            typ_katalogowy = linia.get("typ_katalogowy") or None
            rodzaj = BranchType.LINE.value
            obciazalnosc_a = 0.0
            if typ_katalogowy:
                if typ_katalogowy in katalog.cable_types:
                    rodzaj = BranchType.CABLE.value
                    obciazalnosc_a = float(katalog.cable_types[typ_katalogowy].rated_current_a)
                else:
                    obciazalnosc_a = float(katalog.line_types[typ_katalogowy].rated_current_a)
            elif wymaga_referencji_katalogowej(rodzaj):
                elementy_bez_katalogu.append(linia["id"])

            galezie.append(
                {
                    "ref": linia["id"],
                    "name": linia["typ"],
                    "branch_type": rodzaj,
                    "from_ref": linia["szyna_pocz"],
                    "to_ref": linia["szyna_kon"],
                    "params": {
                        "r_ohm_per_km": linia["R_ohm_km"],
                        "x_ohm_per_km": linia["X_ohm_km"],
                        "b_us_per_km": linia.get("B_uS_km", 0.0),
                        "length_km": linia["długość_km"],
                        "rated_current_a": obciazalnosc_a,
                        "type_ref": typ_katalogowy,
                    },
                }
            )

        napiecia_szyn = {s["id"]: s["napięcie_kV"] for s in szyny}
        for transformator in trafo:
            galezie.append(
                {
                    "ref": transformator["id"],
                    "name": transformator.get("grupa", transformator["id"]),
                    "branch_type": BranchType.TRANSFORMER.value,
                    "from_ref": transformator["szyna_HV"],
                    "to_ref": transformator["szyna_LV"],
                    "params": {
                        "rated_power_mva": transformator["Sn_MVA"],
                        "voltage_hv_kv": napiecia_szyn[transformator["szyna_HV"]],
                        "voltage_lv_kv": napiecia_szyn[transformator["szyna_LV"]],
                        "uk_percent": transformator["uk_pct"],
                        "pk_kw": transformator.get("Pk_kW", 0.0),
                        "vector_group": transformator.get("grupa", "Dyn11"),
                    },
                }
            )

        rekordy_zrodel = [
            {
                "ref": zrodlo["id"],
                "node_ref": zrodlo["szyna"],
                "source_type": "GRID",
                "payload": {
                    "name": zrodlo["id"],
                    # Kanoniczny ksztalt danych zrodla systemowego (jak w kreatorze):
                    # DANE WEJSCIOWE, nie wynik — impedancje liczy warstwa solverowa.
                    "model": "short_circuit_power",
                    "sk3_mva": zrodlo["Sk_MVA"],
                    "rx_ratio": zrodlo["RX_ratio"],
                    "rodzaj_z_arkusza": zrodlo["typ"],
                },
            }
            for zrodlo in zrodla
        ]

        rekordy_odbiorow = [
            {
                "ref": odbior["id"],
                "node_ref": odbior["szyna"],
                "payload": {
                    "name": odbior["id"],
                    "p_mw": odbior["P_MW"],
                    "q_mvar": odbior["Q_Mvar"],
                },
            }
            for odbior in odbiory
        ]

        if elementy_bez_katalogu:
            ostrzezenia.append(
                f"Wymagane domapowanie typu katalogowego dla {len(elementy_bez_katalogu)} "
                f"odcinków — arkusz nie zawierał kolumny 'typ_katalogowy'."
            )

        return (
            SiecZArkusza(
                wezly=wezly,
                galezie=galezie,
                zrodla=rekordy_zrodel,
                odbiory=rekordy_odbiorow,
            ),
            elementy_bez_katalogu,
        )
